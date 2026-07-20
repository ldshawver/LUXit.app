"""CRM contact upsert, source tagging, segment resolution, and SMS recipients."""
from __future__ import annotations

import csv, io, re
from datetime import datetime
from email.utils import parseaddr
from sqlalchemy import or_, func
from extensions import db
from models import Contact, Segment, SegmentMember, SMSRecipient

PHONE_SOURCE_TAG_RULES = [{"phone_number": "+19165989519", "tag": "MyOrder Customer"}]
SYSTEM_KEYWORDS = {"STOP", "STOPALL", "UNSUBSCRIBE", "CANCEL", "END", "QUIT", "HELP", "INFO", "START", "UNSTOP"}


from services.phone_normalization import normalize_phone_e164
from services.contact_intelligence import apply_source_attribution, sync_contact_points

def normalize_phone(number: str | None) -> str:
    return normalize_phone_e164(number)


def _split_tags(tags):
    if not tags:
        return []
    if isinstance(tags, (list, tuple, set)):
        vals = tags
    else:
        vals = re.split(r"[,;|]", str(tags))
    out = []
    for v in vals:
        v = str(v).strip()
        if v and v.lower() not in [x.lower() for x in out]:
            out.append(v)
    return out


def add_contact_tag(contact: Contact, tag: str):
    tags = _split_tags(contact.tags)
    if tag and tag.lower() not in [t.lower() for t in tags]:
        tags.append(tag)
        contact.tags = ", ".join(tags)


def source_tag_rules(config: dict | None = None):
    return (config or {}).get("phone_source_tag_rules") or PHONE_SOURCE_TAG_RULES


def apply_source_tags(contact: Contact, source_phone_number: str | None, config: dict | None = None):
    source_e164 = normalize_phone(source_phone_number)
    for rule in source_tag_rules(config):
        if normalize_phone(rule.get("phone_number")) == source_e164:
            add_contact_tag(contact, rule.get("tag"))


def upsert_contact_from_source(company_id: int, phone: str | None = None, email: str | None = None, *,
                               tenant_id: int | None = None, first_name=None, last_name=None, full_name=None,
                               company=None, tags=None, source_channel="sms", source_phone_number=None,
                               source_provider="twilio", source_context=None, sms_opt_in=None,
                               email_opt_in=None, preserve_opt_out=True) -> Contact:
    now = datetime.utcnow()
    norm = normalize_phone(phone)
    q = Contact.query.filter(Contact.company_id == company_id, Contact.is_active.is_(True))
    contact = None
    if norm:
        contact = q.filter(or_(Contact.normalized_phone == norm, Contact.phone == norm, Contact.phone == phone)).first()
    if not contact and email:
        contact = q.filter(func.lower(Contact.email) == email.strip().lower()).first()
    if not contact:
        contact = Contact(company_id=company_id, tenant_id=tenant_id or company_id, is_active=True, is_subscribed=True, created_at=now)
        db.session.add(contact)
    if not contact.phone and norm:
        contact.phone = norm
    if norm and not contact.normalized_phone:
        contact.normalized_phone = norm
    if email and (not contact.email or "@" not in contact.email):
        contact.email = email.strip()
    if full_name and not contact.name:
        contact.name = full_name.strip()
    if first_name and not contact.first_name:
        contact.first_name = first_name.strip()
    if last_name and not contact.last_name:
        contact.last_name = last_name.strip()
    if company and not contact.company:
        contact.company = company.strip()
    for tag in _split_tags(tags):
        add_contact_tag(contact, tag)
    contact.source_channel = source_channel or contact.source_channel
    contact.source_provider = source_provider or contact.source_provider
    contact.source_context = source_context or contact.source_context
    if source_phone_number:
        contact.source_phone_number = normalize_phone(source_phone_number)
    contact.source = contact.source or source_channel
    contact.source_detail = contact.source_detail or source_context
    source_map = {"sms": "twilio_inbound_sms", "csv_import": "csv_import", "manual": "manual_entry", "api": "api"}
    intel_source = source_map.get((source_channel or "").lower(), source_channel or "unknown")
    # Contact points and source events require a real contact_id. Flush a newly
    # created contact before those child records are constructed.
    db.session.flush()

    sync_contact_points(contact, phone, email, intel_source)
    apply_source_attribution(contact, intel_source, detail=source_context, metadata={"provider": source_provider, "source_phone_number": source_phone_number})
    contact.first_seen_at = contact.first_seen_at or now
    contact.last_seen_at = now
    contact.tenant_id = contact.tenant_id or tenant_id or company_id
    if sms_opt_in is True and not (preserve_opt_out and (contact.sms_opted_out or contact.do_not_sms or contact.sms_opt_out_at)):
        contact.sms_marketing_opt_in = True; contact.sms_consent_status = "opted_in"; contact.sms_marketing_opt_in_at = contact.sms_marketing_opt_in_at or now
    elif sms_opt_in is False:
        contact.sms_marketing_opt_in = False
    if email_opt_in is True and not (preserve_opt_out and (contact.email_unsubscribed or contact.do_not_email)):
        contact.email_opt_in = True; contact.email_subscribed = True; contact.email_unsubscribed = False
    elif email_opt_in is False:
        contact.email_opt_in = False
    apply_source_tags(contact, source_phone_number)
    db.session.flush()
    return contact


def _normalized_label(value):
    """Normalize human-entered tag labels only while resolving their tenant tag ID."""
    return " ".join(str(value or "").split()).casefold()


def _contact_tag_keys(contact):
    return {_normalized_label(tag) for tag in _split_tags(contact.tags)}


def canonical_tag_ids(company_id: int, *, tag_ids=None, tag_names=None, create_missing=False):
    """Resolve tag/segment labels to tenant-owned IDs; foreign IDs are rejected."""
    ids = {int(value) for value in (tag_ids or []) if str(value).isdigit()}
    segments = Segment.query.filter(Segment.company_id == company_id)
    if ids:
        owned = {row.id for row in segments.filter(Segment.id.in_(ids)).all()}
        if owned != ids:
            raise ValueError("One or more selected tag IDs are not available for this company.")
    else:
        owned = set()
    wanted = {_normalized_label(value) for value in (tag_names or []) if _normalized_label(value)}
    if wanted:
        matches = [row for row in segments.all() if _normalized_label(row.name) in wanted]
        found = {_normalized_label(row.name) for row in matches}
        missing = wanted - found
        if missing and create_missing:
            # One-time canonicalization for legacy campaigns/contacts that only
            # stored tag text. All newly persisted audience filters use IDs.
            for key in sorted(missing):
                row = Segment(company_id=company_id, name=" ".join(key.split()), segment_type="contact_tag")
                db.session.add(row); db.session.flush(); matches.append(row)
        elif missing:
            raise ValueError("One or more audience tags do not have a canonical tag ID for this company.")
        owned.update(row.id for row in matches)
    return sorted(owned)


def _campaign_tag_ids(campaign):
    audience = campaign.audience_filter or {}
    raw_ids = audience.get("selected_tag_ids") or audience.get("tag_ids") or []
    names = []
    if not raw_ids and campaign.segment:
        names = _split_tags(campaign.segment)
    return canonical_tag_ids(campaign.company_id, tag_ids=raw_ids, tag_names=names)


def resolve_segment_contacts(company_id: int, segment=None, audience_filter: dict | None = None):
    """Tenant-scoped tag resolution through canonical Segment IDs.

    SegmentMember is authoritative when memberships exist. Legacy contact tag text is
    matched to the *name belonging to the selected tenant Segment ID*, never directly
    to an untrusted campaign label.
    """
    filters = audience_filter or {}
    ids = canonical_tag_ids(
        company_id,
        tag_ids=filters.get("selected_tag_ids") or filters.get("tag_ids") or [],
        tag_names=_split_tags(segment) if segment and not (filters.get("selected_tag_ids") or filters.get("tag_ids")) else [],
        create_missing=True,
    ) if (segment or filters.get("selected_tag_ids") or filters.get("tag_ids")) else []
    contacts = Contact.query.filter(Contact.company_id == company_id).all()
    if not ids:
        return contacts
    segments = Segment.query.filter(Segment.company_id == company_id, Segment.id.in_(ids)).all()
    names = {_normalized_label(row.name) for row in segments}
    member_ids = {row.contact_id for row in SegmentMember.query.filter(
        SegmentMember.segment_id.in_(ids), SegmentMember.is_excluded.is_(False)
    ).all()}
    return [contact for contact in contacts if contact.id in member_ids or bool(_contact_tag_keys(contact) & names)]


def resolve_sms_campaign_recipients(campaign, *, materialize=False):
    """Canonical recipient resolver for preview, sending, scheduling, jobs and reports."""
    matched = resolve_segment_contacts(campaign.company_id, campaign.segment, campaign.audience_filter or {})
    counts = {
        "matching_contacts": len(matched), "contacts_with_phone": 0,
        "unique_phone_numbers": 0, "eligible_recipients": 0,
        "missing_phone_numbers": 0, "invalid_phone_numbers": 0,
        "duplicate_phone_numbers": 0, "opted_out_contacts": 0,
        "missing_sms_consent": 0, "archived_or_suppressed": 0,
    }
    valid_seen = set()
    eligible_seen = set()
    recipients = []
    for contact in matched:
        raw_phone = contact.normalized_phone or contact.phone
        if not str(raw_phone or "").strip():
            counts["missing_phone_numbers"] += 1
            continue
        counts["contacts_with_phone"] += 1
        phone = normalize_phone(raw_phone)
        if not phone:
            counts["invalid_phone_numbers"] += 1
            continue
        if phone in valid_seen:
            counts["duplicate_phone_numbers"] += 1
        valid_seen.add(phone)
        tags = _contact_tag_keys(contact)
        opted_out = bool(contact.sms_opted_out or contact.do_not_sms or contact.sms_opt_out_at or "sms_opt_out" in tags or "no_sms" in tags)
        suppressed = bool(not contact.is_active or contact.archived_at or contact.do_not_market or contact.do_not_contact or contact.status in {"archived", "suppressed", "merged"} or "blocked" in tags)
        consent = bool(contact.sms_marketing_opt_in and contact.sms_consent_status in {"opted_in", "subscribed"})
        if opted_out:
            counts["opted_out_contacts"] += 1
        if not consent:
            counts["missing_sms_consent"] += 1
        if suppressed:
            counts["archived_or_suppressed"] += 1
        if opted_out or suppressed or not consent or phone in eligible_seen:
            continue
        eligible_seen.add(phone)
        recipients.append((contact, phone))
    counts["unique_phone_numbers"] = len(valid_seen)
    counts["eligible_recipients"] = len(recipients)
    # Backward-compatible keys for existing API consumers.
    counts.update(total_matched=counts["matching_contacts"], duplicates_removed=counts["duplicate_phone_numbers"],
                  invalid_numbers_removed=counts["invalid_phone_numbers"], opt_outs_removed=counts["opted_out_contacts"],
                  final_recipients=counts["eligible_recipients"])
    counts["explanation"] = (
        f'{counts["matching_contacts"]} tagged contact(s): {counts["missing_phone_numbers"]} missing phone, '
        f'{counts["invalid_phone_numbers"]} invalid, {counts["duplicate_phone_numbers"]} duplicate, '
        f'{counts["opted_out_contacts"]} opted out, {counts["missing_sms_consent"]} missing consent, '
        f'{counts["archived_or_suppressed"]} archived/suppressed; exactly {counts["eligible_recipients"]} SMS message(s) will be attempted.'
    )
    if materialize:
        SMSRecipient.query.filter_by(company_id=campaign.company_id, campaign_id=campaign.id).delete(synchronize_session=False)
        for contact, phone in recipients:
            db.session.add(SMSRecipient(company_id=campaign.company_id, campaign_id=campaign.id,
                                        contact_id=contact.id, phone_number=phone, status="pending"))
        campaign.estimated_recipient_count = counts["eligible_recipients"]
    return {"counts": counts, "recipients": recipients}


def build_sms_recipient_snapshot(campaign):
    return resolve_sms_campaign_recipients(campaign, materialize=True)["counts"]

HEADER_ALIASES={
 'first_name':['first name','firstname'], 'last_name':['last name','lastname'], 'full_name':['name','full name'], 'company':['company','company name'],
 'email':['email address','email','e-mail address'], 'phone':['phone','mobile phone','cell phone','home phone','work phone','phone 1 - value'],
 'tags':['tags','groups','lists'], 'source':['source','lifecycle stage','lead status'], 'created_at':['create date','created date'], 'notes':['notes']}

def detect_contact_mapping(headers):
    lower={h.strip().lower():h for h in headers}; mapping={}
    for field, aliases in HEADER_ALIASES.items():
        for a in aliases:
            if a in lower: mapping[field]=lower[a]; break
    return mapping


def _parse_bool(v):
    return str(v or '').strip().lower() in {'1','true','yes','y','subscribed','opted in','opt-in'}


def valid_email(email):
    addr = parseaddr(email or '')[1]
    return addr if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', addr or '') else ''


def preview_contact_import(file_bytes: bytes, filename: str):
    name=(filename or '').lower()
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb=load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True); ws=wb.active
        rows=list(ws.iter_rows(values_only=True)); headers=[str(x or '').strip() for x in (rows[0] if rows else [])]
        sample=[dict(zip(headers, r)) for r in rows[1:6]]
    else:
        text=file_bytes.decode('utf-8-sig')
        dialect=csv.excel_tab if name.endswith('.tsv') else csv.Sniffer().sniff(text[:2048], delimiters=',\t;')
        reader=csv.DictReader(io.StringIO(text), dialect=dialect); headers=reader.fieldnames or []; sample=[r for _,r in zip(range(5), reader)]
    return {'headers': headers, 'detected_mapping': detect_contact_mapping(headers), 'sample_rows': sample}


def import_contacts(company_id:int, file_bytes:bytes, filename:str, *, mapping=None, source_provider='generic_csv', imported_list=None, apply_tags=None, sms_subscribed=False, email_subscribed=False, tenant_id=None):
    from models import ContactImportBatch, Segment, SegmentMember
    preview=preview_contact_import(file_bytes, filename); mapping=mapping or preview['detected_mapping']
    batch=ContactImportBatch(company_id=company_id, tenant_id=tenant_id or company_id, source_provider=source_provider, filename=filename, imported_list=imported_list, applied_tags=apply_tags or [], field_mapping=mapping)
    db.session.add(batch); db.session.flush()
    name=(filename or '').lower(); rows=[]
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        ws=load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True).active
        vals=list(ws.iter_rows(values_only=True)); headers=[str(x or '').strip() for x in vals[0]] if vals else []
        rows=[dict(zip(headers, r)) for r in vals[1:]]
    else:
        text=file_bytes.decode('utf-8-sig'); dialect=csv.excel_tab if name.endswith('.tsv') else csv.Sniffer().sniff(text[:2048], delimiters=',\t;')
        rows=list(csv.DictReader(io.StringIO(text), dialect=dialect))
    errors=[]; ok=0
    segment=None
    if imported_list:
        segment=Segment.query.filter_by(company_id=company_id, name=imported_list).first() or Segment(company_id=company_id, name=imported_list, segment_type='imported_list', match_mode='all')
        db.session.add(segment); db.session.flush()
    for idx,row in enumerate(rows, start=2):
        batch.total_rows += 1
        def get(field): return row.get(mapping.get(field,'')) if mapping.get(field) else None
        phone=normalize_phone(get('phone')); email=valid_email(get('email'))
        if not phone and not email:
            errors.append({'row':idx,'error':'missing valid phone or email'}); continue
        try:
            contact=upsert_contact_from_source(company_id, phone or None, email or None, tenant_id=tenant_id or company_id, first_name=get('first_name'), last_name=get('last_name'), full_name=get('full_name'), company=get('company'), tags=_split_tags(get('tags')) + _split_tags(apply_tags), source_channel='csv_import', source_provider=source_provider, source_context='campaign_import', sms_opt_in=_parse_bool(get('sms_opt_in')) or bool(sms_subscribed), email_opt_in=_parse_bool(get('email_opt_in')) or bool(email_subscribed))
            contact.imported_batch_id=batch.id; contact.imported_list=imported_list or contact.imported_list
            if segment and not SegmentMember.query.filter_by(segment_id=segment.id, contact_id=contact.id).first():
                db.session.add(SegmentMember(segment_id=segment.id, contact_id=contact.id, source='import'))
            ok += 1
        except Exception as exc:
            errors.append({'row':idx,'error':str(exc)})
    batch.success_count=ok; batch.failure_count=len(errors); batch.error_report=errors
    db.session.flush()
    return {'batch': batch, 'success_count': ok, 'failure_count': len(errors), 'errors': errors, 'mapping': mapping}
